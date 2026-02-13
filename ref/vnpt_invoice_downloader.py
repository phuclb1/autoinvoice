#!/usr/bin/env python3
"""
VNPT Invoice Downloader
Tự động tìm kiếm và download hóa đơn từ VNPT Invoice portal
Sử dụng Playwright và Claude API để giải captcha
"""

import asyncio
import os
import sys
import argparse
from pathlib import Path
from typing import Optional, List

from playwright.async_api import async_playwright, Browser, Page, BrowserContext
from google import genai
from google.genai import types
from openpyxl import load_workbook
from openai import OpenAI
import base64


class VNPTInvoiceDownloader:
    """Lớp tự động hóa tìm kiếm và download hóa đơn từ VNPT"""

    def __init__(
        self,
        invoice_code: str,
        download_dir: str = "./downloads",
        headless: bool = False,
        claude_api_key: Optional[str] = None,
        ai_provider: str = "gemini"
    ):
        """
        Khởi tạo downloader

        Args:
            invoice_code: Mã hóa đơn cần tìm
            download_dir: Thư mục lưu file download
            headless: Chạy ẩn danh (không hiển thị browser)
            claude_api_key: API key cho AI provider (Gemini hoặc OpenAI) để giải captcha
            ai_provider: Loại AI provider để giải captcha ('gemini' hoặc 'openai')
        """
        self.invoice_code = invoice_code
        self.download_dir = Path(download_dir)
        self.download_dir.mkdir(parents=True, exist_ok=True)
        self.headless = headless
        self.ai_provider = ai_provider.lower()
        
        # API key setup based on provider
        if self.ai_provider == "openai":
            self.claude_api_key = claude_api_key or os.getenv("OPENAI_API_KEY")
        else:  # default to gemini
            self.claude_api_key = claude_api_key or os.getenv("GEMINI_API_KEY")
            
        self.browser: Optional[Browser] = None
        self.context: Optional[BrowserContext] = None
        self.page: Optional[Page] = None

        # URL trang tìm kiếm
        self.url = os.getenv("INVOICE_URL", "https://3701642642-010-tt78.vnpt-invoice.com.vn/HomeNoLogin/SearchByFkey")

    async def _setup_browser(self):
        """Cấu hình và khởi tạo Playwright browser"""
        self.playwright = await async_playwright().start()

        # Cấu hình download
        self.browser = await self.playwright.chromium.launch(
            headless=self.headless,
            args=['--no-sandbox', '--disable-setuid-sandbox']
        )

        self.context = await self.browser.new_context(
            accept_downloads=True,
            viewport={'width': 1920, 'height': 1080}
        )

        self.page = await self.context.new_page()

        # Cài đặt default download behavior
        await self.page.route("**/*", lambda route: route.continue_())

    async def _solve_captcha_with_gemini(self, screenshot_bytes: bytes) -> str:
        """
        Giải captcha sử dụng Google Gemini API

        Args:
            screenshot_bytes: Ảnh captcha dưới dạng bytes

        Returns:
            Chuỗi captcha đã giải
        """
        if not self.claude_api_key:
            raise Exception("Không có Gemini API key!")

        try:
            print("  - Calling Gemini API...")
            print(f"  - API Key: {self.claude_api_key[:20]}...")
            print(f"  - Image size: {len(screenshot_bytes)} bytes")

            # Khởi tạo Gemini client với API key
            client = genai.Client(api_key=self.claude_api_key)

            # Prompt để giải captcha
            prompt = """Please extract the text from this captcha image.
Return ONLY the captcha text, nothing else. No explanations, no quotes, just the raw text.
The captcha usually contains 4 alphanumeric characters."""

            # Gọi Gemini API
            response = client.models.generate_content(
                model='gemini-3-flash-preview',
                contents=[
                    types.Part.from_bytes(
                        data=screenshot_bytes,
                        mime_type='image/png',
                    ),
                    prompt
                ]
            )

            print(f"  - Response received")
            print(f"  - Response text: '{response.text}'")

            captcha_text = response.text.strip()
            print(f"✓ Gemini đã giải captcha: {captcha_text}")
            return captcha_text

        except Exception as e:
            print(f"Lỗi khi gọi Gemini API: {e}")
            import traceback
            traceback.print_exc()
            raise

    async def _solve_captcha_with_openai(self, screenshot_bytes: bytes) -> str:
        """
        Giải captcha sử dụng OpenAI API (GPT-4o-mini)

        Args:
            screenshot_bytes: Ảnh captcha dưới dạng bytes

        Returns:
            Chuỗi captcha đã giải
        """
        if not self.claude_api_key:
            raise Exception("Không có OpenAI API key!")

        try:
            print("  - Calling OpenAI API (GPT-4o-mini)...")
            print(f"  - API Key: {self.claude_api_key[:20]}...")
            print(f"  - Image size: {len(screenshot_bytes)} bytes")

            # Khởi tạo OpenAI client với API key
            client = OpenAI(api_key=self.claude_api_key)

            # Encode image to base64
            base64_image = base64.b64encode(screenshot_bytes).decode('utf-8')

            # Prompt để giải captcha
            prompt = """Please extract the text from this captcha image.
Return ONLY the captcha text, nothing else. No explanations, no quotes, just the raw text.
The captcha usually contains 4 alphanumeric characters."""

            # Gọi OpenAI API
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": prompt
                            },
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/png;base64,{base64_image}"
                                }
                            }
                        ]
                    }
                ],
                max_tokens=100
            )

            print(f"  - Response received")
            captcha_text = response.choices[0].message.content.strip()
            print(f"  - Response text: '{captcha_text}'")
            print(f"✓ OpenAI GPT-4o-mini đã giải captcha: {captcha_text}")
            return captcha_text

        except Exception as e:
            print(f"Lỗi khi gọi OpenAI API: {e}")
            import traceback
            traceback.print_exc()
            raise

    async def _solve_captcha_manual(self) -> str:
        """
        Giải captcha theo cách thủ công
        Hiển thị ảnh captcha và yêu cầu người dùng nhập
        """
        try:
            # Tìm ảnh captcha
            captcha_element = await self.page.wait_for_selector(
                "img[src*='captcha']",
                timeout=10000
            )

            # Lưu ảnh captcha
            captcha_path = self.download_dir / "captcha_temp.png"
            await captcha_element.screenshot(path=str(captcha_path))

            print(f"\n{'='*50}")
            print(f"Captcha đã được lưu tại: {captcha_path}")
            print(f"{'='*50}")

            # Mở ảnh captcha (tùy OS)
            if sys.platform == "darwin":
                os.system(f"open {captcha_path}")
            else:
                os.system(f"start {captcha_path}")

            # Nhập captcha từ bàn phím
            loop = asyncio.get_event_loop()
            captcha_text = await loop.run_in_executor(
                None,
                lambda: input("Nhập mã xác thực (captcha): ").strip()
            )

            return captcha_text

        except Exception as e:
            print(f"Không tìm thấy ảnh captcha: {e}")
            return ""

    async def _input_captcha(self) -> bool:
        """Nhập mã captcha với cơ chế retry"""
        print("\n{'='*50}")
        print("BẮT ĐẦU QUY TRÌNH GIẢI CAPTCHA")
        print(f"{'='*50}\n")

        max_attempts = 3
        use_manual = False

        for attempt in range(max_attempts):
            try:
                if attempt > 0:
                    print(f"\n🔄 Thử lại lần {attempt + 1}/{max_attempts}...")
                    # Reload lại check nếu cần, hoặc captcha tự refresh sau khi submit sai

                # Step 1: Tìm ảnh captcha
                print("Step 1: Tìm ảnh captcha...")
                # Chờ một chút để đảm bảo ảnh đã load (đặc biệt là sau khi reload)
                await asyncio.sleep(1)
                
                try:
                    captcha_element = await self.page.wait_for_selector('form img[src="/Captcha/Show"]', timeout=5000)
                except:
                    # Fallback selector
                    captcha_element = await self.page.wait_for_selector("img[src*='captcha']", timeout=5000)
                    
                print("✓ Đã tìm thấy ảnh captcha")

                # Step 2: Download ảnh captcha
                print("\nStep 2: Download ảnh captcha...")
                captcha_bytes = await captcha_element.screenshot()

                # Lưu ảnh ra file để debug
                captcha_debug_path = self.download_dir / f"debug_captcha_attempt_{attempt+1}.png"
                with open(captcha_debug_path, "wb") as f:
                    f.write(captcha_bytes)
                print(f"✓ Đã lưu ảnh captcha tại: {captcha_debug_path}")

                # Step 3: Giải captcha
                captcha_text = ""
                
                if not use_manual:
                    # Thử dùng AI trước
                    if self.ai_provider == "openai":
                        print("\nStep 3: Giải captcha bằng OpenAI GPT-4o-mini...")
                    else:
                        print("\nStep 3: Giải captcha bằng Gemini 2.0 Flash...")
                        
                    if self.claude_api_key:
                        try:
                            if self.ai_provider == "openai":
                                captcha_text = await self._solve_captcha_with_openai(captcha_bytes)
                            else:
                                captcha_text = await self._solve_captcha_with_gemini(captcha_bytes)
                        except Exception as e:
                            print(f"✗ Không thể giải captcha bằng {self.ai_provider.upper()}: {e}")
                            print("Chuyển sang chế độ manual...")
                            use_manual = True
                    else:
                        print("Không có API key, dùng chế độ manual")
                        use_manual = True
                
                # Nếu use_manual được bật (hoặc vừa bật do lỗi AI)
                if use_manual:
                     # Mở ảnh để người dùng xem
                    if sys.platform == "darwin":
                        os.system(f"open {captcha_debug_path}")
                    else:
                        os.system(f"start {captcha_debug_path}")
                        
                    print(f"\n{'='*50}")
                    print("⌨ VUI LÒNG NHẬP CAPTCHA THỦ CÔNG")
                    print(f"{'='*50}")
                    
                    loop = asyncio.get_event_loop()
                    captcha_text = await loop.run_in_executor(
                        None,
                        lambda: input("Nhập mã captcha từ ảnh: ").strip()
                    )

                if not captcha_text:
                    print("✗ Không có text captcha!")
                    use_manual = True # Force manual next time
                    continue

                # Step 4: Nhập captcha vào form
                print(f"\nStep 4: Nhập captcha vào form...")
                print(f"  - Captcha text: '{captcha_text}'")

                # Clear cũ và nhập mới
                await self.page.fill(".captcha_input.form-control", "")
                await self.page.fill(".captcha_input.form-control", captcha_text)
                print(f"✓ Đã nhập captcha: {captcha_text}")

                # Step 5: Click button submit
                print(f"\nStep 5: Click button tìm kiếm...")
                await self.page.click("button[type='submit']")
                print(f"✓ Đã click button submit")

                # Chờ load kết quả
                await self.page.wait_for_load_state("networkidle", timeout=15000)
                await asyncio.sleep(2)

                # KIỂM TRA LỖI SAU KHI SUBMIT
                # Kiểm tra xem có alert lỗi hay không
                # Thông thường VNPT Invoice báo lỗi bằng alert đỏ hoặc text
                error_element = await self.page.query_selector(".validation-summary-errors, .alert-danger, label.error")
                error_text = ""
                if error_element:
                    error_text = await error_element.text_content()
                
                # Hoặc kiểm tra xem URL có thay đổi không, hoặc form captcha còn đó không
                # Nếu form captcha vẫn còn và có dòng thông báo lỗi
                if error_text and ("sai" in error_text.lower() or "không đúng" in error_text.lower()):
                    print(f"⚠ LỖI TỪ WEBSITE: {error_text.strip()}")
                    print("👉 Captcha không chính xác, thử lại với manual input...")
                    use_manual = True
                    
                    # Refresh captcha nếu cần (thường click vào ảnh)
                    # await captcha_element.click() 
                    # await asyncio.sleep(1)
                    continue
                
                # Nếu không thấy lỗi rõ ràng, check xem đã vào được trang kết quả chưa
                # Trang kết quả thường có bảng hoặc thông tin hóa đơn
                # Nếu vẫn còn nút "Tìm kiếm" và không có bảng kết quả -> Khả năng cao là fail
                # Nhưng an toàn nhất là return True để _download_invoice check tiếp
                # Nếu _download_invoice fail, nó sẽ gọi _retry_with_manual_captcha
                
                print(f"\n{'='*50}")
                print("✓ Đã submit captcha (không phát hiện lỗi ngay lập tức)")
                print(f"{'='*50}\n")

                return True

            except Exception as e:
                print(f"\n✗ Lỗi khi nhập captcha (attempt {attempt+1}): {e}")
                import traceback
                traceback.print_exc()
                use_manual = True # Switch to manual on crash
                
                # Thử refresh trang để reset trạng thái
                try:
                    await self.page.reload()
                    await self._input_invoice_code() # Nhập lại invoice code sau reload
                except:
                    pass

        print("✗ Đã hết số lần thử giải captcha!")
        return False

    async def _input_invoice_code(self):
        """Nhập mã hóa đơn - sử dụng getByRole API"""
        try:
            # Dùng getByRole để tìm textbox chính xác theo label
            await self.page.get_by_role("textbox", name="Nhập mã tra cứu hóa đơn").fill(self.invoice_code)
            print(f"✓ Đã nhập mã hóa đơn: {self.invoice_code}")

        except Exception as e:
            print(f"Lỗi khi nhập mã hóa đơn: {e}")
            # Fallback: thử các selector khác
            try:
                await self.page.fill("#Fkey", self.invoice_code)
                print(f"✓ Đã nhập mã hóa đơn (fallback): {self.invoice_code}")
            except Exception:
                raise Exception("Không tìm thấy ô nhập mã hóa đơn!")

    async def _submit_search(self):
        """Click nút tìm kiếm - sử dụng getByRole API"""
        try:
            # Dùng getByRole để tìm nút button chính xác
            await self.page.get_by_role("button", name=" Tìm kiếm").click()
            print("✓ Đã click nút tìm kiếm")

            # Chờ load kết quả
            await self.page.wait_for_load_state("networkidle", timeout=15000)
            await asyncio.sleep(2)

        except Exception as e:
            # Fallback: thử các selector khác
            try:
                await self.page.click("button[type='submit'], input[type='submit']")
                print("✓ Đã click nút tìm kiếm (fallback)")
                await self.page.wait_for_load_state("networkidle", timeout=15000)
                await asyncio.sleep(2)
            except Exception:
                raise Exception(f"Không tìm thấy nút tìm kiếm: {e}")

    async def _download_invoice(self) -> bool:
        """Download file hóa đơn với retry nếu lỗi"""
        try:
            print("\n{'='*50}")
            print("BẮT ĐẦU DOWNLOAD FILE")
            print(f"{'='*50}\n")

            # Tìm link download theo title="Tải file pdf"
            print("Step 1: Tìm link download...")
            download_link = await self.page.query_selector("a[title='Tải file pdf'][href*='/HomeNoLogin/downloadPDF']")

            if download_link:
                print("✓ Tìm thấy link download PDF")

                # Lấy href để debug
                href = await download_link.get_attribute("href")
                print(f"  - HREF: {href}")

                print("Step 2: Click download...")
                async with self.page.expect_download(timeout=30000) as download_info:
                    await download_link.click()

                download = await download_info.value

                # Lưu file
                download_path = self.download_dir / download.suggested_filename
                await download.save_as(download_path)

                print(f"✓ File đã được download: {download_path.name}")
                print(f"✓ Đường dẫn: {download_path.absolute()}")

                print(f"\n{'='*50}")
                print("✓ DOWNLOAD THÀNH CÔNG!")
                print(f"{'='*50}\n")

                return True
            else:
                # Fallback: Tìm link PDF theo title hoặc href
                print("Không tìm thấy link chính xác, thử các cách khác...")

                # Thử tìm theo title
                download_link = await self.page.query_selector("a[title='Tải file pdf']")
                if download_link:
                    print("✓ Tìm thấy theo title='Tải file pdf'")
                else:
                    # Thử tìm theo href
                    download_links = await self.page.query_selector_all("a[href*='/HomeNoLogin/downloadPDF']")
                    if download_links:
                        print(f"✓ Tìm thấy {len(download_links)} link downloadPDF")
                        download_link = download_links[0]
                    else:
                        # Fallback cuối cùng
                        download_links = await self.page.query_selector_all("a[href*='.pdf'], a[download]")
                        if not download_links:
                            print("✗ Không tìm thấy link download PDF!")
                            return False
                        download_link = download_links[0]

                async with self.page.expect_download(timeout=30000) as download_info:
                    await download_link.click()

                download = await download_info.value
                download_path = self.download_dir / download.suggested_filename
                await download.save_as(download_path)

                print(f"✓ File đã được download: {download_path.name}")

                return True

        except Exception as e:
            print(f"\n✗ Lỗi khi download: {e}")
            import traceback
            traceback.print_exc()
            
            # Kiểm tra xem có phải lỗi captcha không
            print("\n⚠ Download thất bại, có thể do lỗi captcha hoặc session timeout")
            print("Sẽ thử lại với captcha manual...")
            
            return False

    async def _retry_with_manual_captcha(self) -> bool:
        """Retry download với captcha manual khi gặp lỗi"""
        try:
            print("\n{'='*50}")
            print("THỬ LẠI VỚI CAPTCHA MANUAL")
            print(f"{'='*50}\n")
            
            # Lưu ảnh màn hình hiện tại để debug
            error_screenshot = self.download_dir / "error_screenshot.png"
            await self.page.screenshot(path=str(error_screenshot))
            print(f"📸 Đã lưu screenshot lỗi: {error_screenshot}")
            
            # Mở ảnh để người dùng xem
            if sys.platform == "darwin":
                os.system(f"open {error_screenshot}")
            else:
                os.system(f"start {error_screenshot}")
            
            # Kiểm tra xem có form captcha không
            captcha_element = await self.page.query_selector("img[src*='captcha'], img[src='/Captcha/Show']")
            
            if captcha_element:
                print("\n✓ Phát hiện form captcha, yêu cầu nhập lại...")
                
                # Lưu ảnh captcha
                captcha_path = self.download_dir / "captcha_retry.png"
                await captcha_element.screenshot(path=str(captcha_path))
                print(f"📸 Đã lưu ảnh captcha: {captcha_path}")
                
                # Mở ảnh captcha
                if sys.platform == "darwin":
                    os.system(f"open {captcha_path}")
                else:
                    os.system(f"start {captcha_path}")
                
                # Yêu cầu người dùng nhập captcha
                print(f"\n{'='*50}")
                print("⌨ VUI LÒNG NHẬP CAPTCHA THỦ CÔNG")
                print(f"{'='*50}")
                
                loop = asyncio.get_event_loop()
                captcha_text = await loop.run_in_executor(
                    None,
                    lambda: input("Nhập mã captcha từ ảnh: ").strip()
                )
                
                if not captcha_text:
                    print("✗ Không nhận được mã captcha!")
                    return False
                
                # Nhập captcha vào form
                await self.page.fill(".captcha_input.form-control", captcha_text)
                print(f"✓ Đã nhập captcha: {captcha_text}")
                
                # Click nút tìm kiếm/submit
                await self.page.click("button[type='submit']")
                print("✓ Đã click nút submit")
                
                # Chờ load kết quả
                await self.page.wait_for_load_state("networkidle", timeout=15000)
                await asyncio.sleep(2)
                
                # Thử download lại
                return await self._download_invoice()
                
            else:
                print("\n⚠ Không tìm thấy form captcha")
                print("Có thể lỗi do:")
                print("  - Session timeout")
                print("  - Hóa đơn không tồn tại")
                print("  - Website bị lỗi")
                
                # Hỏi người dùng có muốn thử lại không
                print(f"\n{'='*50}")
                loop = asyncio.get_event_loop()
                retry = await loop.run_in_executor(
                    None,
                    lambda: input("Bạn có muốn thử lại? (y/n): ").strip().lower()
                )
                
                if retry == 'y':
                    # Reload trang và thử lại từ đầu
                    print("\n🔄 Đang reload trang...")
                    await self.page.goto(self.url, wait_until="networkidle")
                    await asyncio.sleep(2)
                    
                    # Nhập lại mã hóa đơn
                    await self._input_invoice_code()
                    
                    # Nhập captcha manual
                    if await self._input_captcha():
                        # Thử download
                        return await self._download_invoice()
                
                return False
                
        except Exception as e:
            print(f"\n✗ Lỗi khi retry: {e}")
            import traceback
            traceback.print_exc()
            return False

    async def run(self) -> bool:
        """
        Chạy quy trình tìm kiếm và download

        Returns:
            True nếu thành công, False nếu thất bại
        """
        try:
            print(f"\n{'='*50}")
            print("VNPT INVOICE DOWNLOADER")
            print(f"{'='*50}")
            print(f"Mã hóa đơn: {self.invoice_code}")
            print(f"Thư mục download: {self.download_dir}")
            print(f"URL: {self.url}")
            print(f"AI Provider: {self.ai_provider.upper()}")
            print(f"AI API: {'✓' if self.claude_api_key else '✗ (sẽ dùng manual)'}")
            print(f"{'='*50}\n")

            # Khởi tạo browser
            await self._setup_browser()

            # Mở trang web
            print("⏳ Đang mở trang web...")
            await self.page.goto(self.url, wait_until="networkidle")
            await asyncio.sleep(2)

            # Nhập mã hóa đơn
            await self._input_invoice_code()

            # Nhập captcha và submit (bao gồm cả click button submit)
            if not await self._input_captcha():
                return False

            # Download file
            success = await self._download_invoice()

            # Nếu download thất bại, thử lại với captcha manual
            if not success:
                print(f"\n{'='*50}")
                print("⚠ DOWNLOAD LẦN ĐẦU THẤT BẠI - THỬ LẠI")
                print(f"{'='*50}\n")
                
                success = await self._retry_with_manual_captcha()

            if success:
                print(f"\n{'='*50}")
                print("✓ DOWNLOAD THÀNH CÔNG!")
                print(f"{'='*50}\n")
            else:
                print(f"\n{'='*50}")
                print("✗ DOWNLOAD THẤT BẠI!")
                print(f"{'='*50}\n")

            return success

        except Exception as e:
            print(f"\n✗ Lỗi: {e}")
            import traceback
            traceback.print_exc()
            return False

        finally:
            # Đóng browser
            if self.browser:
                await self.browser.close()
            await self.playwright.stop()


def read_invoice_codes_from_excel(file_path: str) -> List[str]:
    """
    Đọc danh sách mã tra cứu hóa đơn từ file Excel
    
    Args:
        file_path: Đường dẫn đến file Excel
        
    Returns:
        List các mã tra cứu hóa đơn
    """
    try:
        print(f"\n📄 Đang đọc file Excel: {file_path}")
        wb = load_workbook(file_path)
        ws = wb.active
        
        invoice_codes = []
        header_row = None
        invoice_code_col = None
        
        # Tìm header row và cột "MÃ TRA CỨU HÓA ĐƠN ĐIỆN TỬ"
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            for col_idx, cell in enumerate(row):
                if cell and isinstance(cell, str) and 'MÃ TRA CỨU' in cell.upper():
                    header_row = row_idx
                    invoice_code_col = col_idx
                    print(f"✓ Tìm thấy cột 'MÃ TRA CỨU' ở row {header_row}, column {col_idx + 1}")
                    break
            if header_row:
                break
        
        if not header_row or invoice_code_col is None:
            raise Exception("Không tìm thấy cột 'MÃ TRA CỨU HÓA ĐƠN ĐIỆN TỬ' trong file Excel!")
        
        # Đọc các mã tra cứu từ cột đã tìm thấy
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx <= header_row:  # Bỏ qua header
                continue
            
            invoice_code = row[invoice_code_col] if invoice_code_col < len(row) else None
            
            # Chỉ lấy các giá trị không rỗng và có pattern hợp lệ
            if invoice_code and str(invoice_code).strip():
                code = str(invoice_code).strip()
                
                # Filter: Chỉ lấy mã có pattern CXXTLK (ví dụ: C25TLK0019654_Ln)
                # Bỏ qua các dòng như header tiếng Anh hoặc chữ ký
                if 'C' in code and '_' in code:
                    invoice_codes.append(code)
        
        print(f"✓ Đã đọc được {len(invoice_codes)} mã tra cứu hóa đơn hợp lệ")
        
        # Hiển thị preview 5 mã đầu tiên
        if invoice_codes:
            print(f"\n📋 Preview {min(5, len(invoice_codes))} mã đầu tiên:")
            for i, code in enumerate(invoice_codes[:5], 1):
                print(f"  {i}. {code}")
            if len(invoice_codes) > 5:
                print(f"  ... và {len(invoice_codes) - 5} mã khác")
        
        return invoice_codes
        
    except Exception as e:
        print(f"✗ Lỗi khi đọc file Excel: {e}")
        import traceback
        traceback.print_exc()
        return []


async def main():
    """Hàm main - hỗ trợ download đơn lẻ hoặc batch từ Excel"""
    # Parse command line arguments
    parser = argparse.ArgumentParser(
        description='VNPT Invoice Downloader - Tự động download hóa đơn từ VNPT (download từng mã 1)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Download 1 hóa đơn:
  python vnpt_invoice_downloader.py --code C25TLK0019654_Ln
  
  # Download batch từ Excel (sẽ download từng mã 1):
  python vnpt_invoice_downloader.py --excel sample.xlsx
  
  # Download với browser hiển thị:
  python vnpt_invoice_downloader.py --excel sample.xlsx --show-browser
  
  # Download với Gemini API key:
  python vnpt_invoice_downloader.py --code C25TLK0019654_Ln --api-key YOUR_API_KEY
        """
    )
    
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument(
        '--code', '-c',
        type=str,
        help='Mã tra cứu hóa đơn đơn lẻ'
    )
    group.add_argument(
        '--excel', '-e',
        type=str,
        help='Đường dẫn file Excel chứa danh sách mã tra cứu (sẽ download từng mã 1)'
    )
    
    parser.add_argument(
        '--download-dir', '-d',
        type=str,
        default='./vnpt_invoices',
        help='Thư mục lưu file download (default: ./vnpt_invoices)'
    )
    
    parser.add_argument(
        '--show-browser',
        action='store_true',
        help='Hiển thị browser (default: chạy ẩn danh)'
    )
    
    parser.add_argument(
        '--api-key',
        type=str,
        help='AI API key (mặc định đọc từ GEMINI_API_KEY hoặc OPENAI_API_KEY env var)'
    )
    
    parser.add_argument(
        '--ai-provider',
        type=str,
        choices=['gemini', 'openai'],
        default='gemini',
        help='AI provider để giải captcha: gemini hoặc openai (default: gemini)'
    )
    
    args = parser.parse_args()
    
    # API key cho AI provider
    if args.ai_provider == 'openai':
        ai_api_key = args.api_key or os.getenv("OPENAI_API_KEY")
        api_key_name = "OPENAI_API_KEY"
    else:
        ai_api_key = args.api_key or os.getenv("GEMINI_API_KEY")
        api_key_name = "GEMINI_API_KEY"
    
    if not ai_api_key:
        print(f"⚠ CẢNH BÁO: Không tìm thấy {api_key_name}!")
        print("  - Sẽ dùng chế độ nhập captcha thủ công")
        print(f"  - Để dùng {args.ai_provider.upper()} API, set: export {api_key_name}='your-api-key'")
        print()
    
    # Lấy danh sách mã tra cứu
    if args.code:
        # Single invoice
        invoice_codes = [args.code]
        print(f"📌 Mode: Download đơn lẻ")
    else:
        # Batch from Excel
        print(f"📌 Mode: Download batch từ Excel (từng mã 1)")
        invoice_codes = read_invoice_codes_from_excel(args.excel)
        
        if not invoice_codes:
            print("✗ Không có mã tra cứu nào để download!")
            return 1
    
    # Summary
    print(f"\n{'='*60}")
    print(f"VNPT INVOICE DOWNLOADER")
    print(f"{'='*60}")
    print(f"📊 Số lượng hóa đơn: {len(invoice_codes)}")
    print(f"📁 Thư mục lưu: {args.download_dir}")
    print(f"🤖 AI Provider: {args.ai_provider.upper()}")
    print(f"🔑 AI API: {'✓ Enabled' if ai_api_key else '✗ Disabled (manual mode)'}")
    print(f"👁  Browser mode: {'Visible' if args.show_browser else 'Headless (ẩn)'}")
    print(f"{'='*60}\n")
    
    # Download từng hóa đơn
    success_count = 0
    failed_count = 0
    failed_codes = []
    
    for idx, invoice_code in enumerate(invoice_codes, 1):
        print(f"\n{'#'*60}")
        print(f"📥 [{idx}/{len(invoice_codes)}] Đang download: {invoice_code}")
        print(f"{'#'*60}\n")
        
        try:
            downloader = VNPTInvoiceDownloader(
                invoice_code=invoice_code,
                download_dir=args.download_dir,
                headless=not args.show_browser,
                claude_api_key=ai_api_key,
                ai_provider=args.ai_provider
            )
            
            success = await downloader.run()
            
            if success:
                success_count += 1
                print(f"✅ [{idx}/{len(invoice_codes)}] Thành công: {invoice_code}")
            else:
                failed_count += 1
                failed_codes.append(invoice_code)
                print(f"❌ [{idx}/{len(invoice_codes)}] Thất bại: {invoice_code}")
            
            # Delay giữa các lần download để tránh bị block
            if idx < len(invoice_codes):
                wait_time = 2
                print(f"\n⏳ Chờ {wait_time}s trước khi download tiếp...")
                await asyncio.sleep(wait_time)
                
        except Exception as e:
            failed_count += 1
            failed_codes.append(invoice_code)
            print(f"❌ [{idx}/{len(invoice_codes)}] Lỗi: {invoice_code} - {e}")
    
    # Final summary
    print(f"\n{'='*60}")
    print(f"📊 KẾT QUẢ TỔNG KẾT")
    print(f"{'='*60}")
    print(f"✅ Thành công: {success_count}/{len(invoice_codes)}")
    print(f"❌ Thất bại: {failed_count}/{len(invoice_codes)}")
    
    if failed_codes:
        print(f"\n❌ Danh sách mã thất bại:")
        for code in failed_codes:
            print(f"   - {code}")
    
    print(f"{'='*60}\n")
    
    return 0 if failed_count == 0 else 1


if __name__ == "__main__":
    exit(asyncio.run(main()))
